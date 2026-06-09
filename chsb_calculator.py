#!/usr/bin/env python3
"""
CHSB Ball Hisoblash Dasturi
============================
Excel fayllardan CHSB ustunidagi ballarni o'qib,
har bir o'quvchi uchun yig'indi va o'rtachani hisoblab chiqaradi.

Ishlatish:
    python chsb_calculator.py fayl1.xlsx fayl2.xlsx ...
    yoki
    python chsb_calculator.py  (joriy papkadagi barcha .xlsx fayllarni o'qiydi)
"""

import sys
import os
import glob
import pandas as pd
import openpyxl
from openpyxl import load_workbook


def find_chsb_column(ws):
    """
    CHSB ustunini topadi.
    Sarlavha qatorini qidirib, 'CHSB' so'zini o'z ichiga olgan ustunni topadi.
    """
    for row in ws.iter_rows(min_row=1, max_row=10):
        for cell in row:
            val = str(cell.value) if cell.value is not None else ""
            if "CHSB" in val.upper():
                return cell.column, cell.row
    return None, None


def find_student_name_column(ws, header_row):
    """
    O'quvchi ismi ustunini topadi (odatda 'Son' yoki ism/familiya sarlavhasi).
    """
    for cell in ws[header_row]:
        val = str(cell.value) if cell.value is not None else ""
        if any(kw in val.lower() for kw in ["ism", "familiya", "son", "o'quvchi", "oquvchi", "name"]):
            return cell.column
    # topilmasa, 2-ustunni qaytaradi (odatda ism ustuni)
    return 2


def is_numeric(value):
    """Qiymat raqammi yoki yo'qligini tekshiradi."""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip())
        return True
    except (ValueError, TypeError):
        return False


def process_excel_file(filepath):
    """
    Bitta Excel faylini qayta ishlaydi.
    Barcha varaqlarni ko'rib chiqadi va CHSB ustunini topadi.
    """
    print(f"\n{'='*60}")
    print(f"Fayl: {os.path.basename(filepath)}")
    print(f"{'='*60}")

    try:
        wb = load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"  [XATO] Faylni ochib bo'lmadi: {e}")
        return None

    all_results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n  Varaq: '{sheet_name}'")

        # CHSB ustunini top
        chsb_col, header_row = find_chsb_column(ws)

        if chsb_col is None:
            print("  [!] CHSB ustuni topilmadi, o'tkazib yuborildi.")
            continue

        print(f"  CHSB ustuni: {chsb_col}-ustun, sarlavha qatori: {header_row}")

        # Ism ustunini top
        name_col = find_student_name_column(ws, header_row)

        students = []
        chsb_values = []

        # Ma'lumot qatorlarini o'qi (sarlavhadan keyin)
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            # Bo'sh qatorni o'tkazib yubor
            if all(cell is None for cell in row):
                continue

            # Qator raqami (tartib raqami) — 1-ustunda
            row_num = row[0]
            # Ism — name_col - 1 (0-indexed)
            name = row[name_col - 1] if len(row) >= name_col else None
            # CHSB bali — chsb_col - 1 (0-indexed)
            chsb_val = row[chsb_col - 1] if len(row) >= chsb_col else None

            # Tartib raqami raqam bo'lmasa (sarlavha, bo'sh, izoh) — o'tkazib yubor
            if not is_numeric(row_num) and not name:
                continue

            student_name = str(name).strip() if name else f"O'quvchi #{row_num}"

            if is_numeric(chsb_val):
                val = float(chsb_val)
                students.append(student_name)
                chsb_values.append(val)
            else:
                # Raqam bo'lmagan (s, k, dq, None) — nol sifatida hisoblanmaydi
                students.append(student_name)
                chsb_values.append(None)

        if not students:
            print("  [!] O'quvchilar topilmadi.")
            continue

        # Natijalarni chiqar
        col1 = "O'quvchi ismi"
        print(f"\n  {'№':<4} {col1:<30} {'CHSB bali':>10}")
        print(f"  {'-'*4} {'-'*30} {'-'*10}")

        valid_scores = []
        for i, (name, score) in enumerate(zip(students, chsb_values), 1):
            score_str = f"{score:.1f}" if score is not None else "—"
            print(f"  {i:<4} {name:<30} {score_str:>10}")
            if score is not None:
                valid_scores.append(score)

        print(f"\n  {'─'*46}")
        if valid_scores:
            total = sum(valid_scores)
            average = total / len(valid_scores)
            print(f"  Jami o'quvchilar soni    : {len(students)}")
            print(f"  Ball kiritilgan o'quvchilar: {len(valid_scores)}")
            print(f"  Yig'indi (jami ball)      : {total:.1f}")
            print(f"  O'rtacha ball             : {average:.2f}")
            print(f"  Eng yuqori ball           : {max(valid_scores):.1f}")
            print(f"  Eng past ball             : {min(valid_scores):.1f}")
        else:
            print("  [!] Hisoblash uchun raqamli ball topilmadi.")

        sheet_result = {
            "fayl": os.path.basename(filepath),
            "varaq": sheet_name,
            "jami_oquvchi": len(students),
            "ball_kiritilgan": len(valid_scores),
            "yigindi": sum(valid_scores) if valid_scores else 0,
            "ortacha": sum(valid_scores) / len(valid_scores) if valid_scores else 0,
            "max": max(valid_scores) if valid_scores else 0,
            "min": min(valid_scores) if valid_scores else 0,
        }
        all_results.append(sheet_result)

    wb.close()
    return all_results


def print_summary(all_file_results):
    """Barcha fayllar bo'yicha umumiy xulosani chiqaradi."""
    print(f"\n{'='*60}")
    print("  UMUMIY XULOSA (barcha fayllar)")
    print(f"{'='*60}")

    all_scores_sum = 0
    all_scores_count = 0

    for result in all_file_results:
        if result:
            for sheet in result:
                all_scores_sum += sheet["yigindi"]
                all_scores_count += sheet["ball_kiritilgan"]

    if all_scores_count > 0:
        grand_average = all_scores_sum / all_scores_count
        print(f"  Barcha fayllar bo'yicha jami ball   : {all_scores_sum:.1f}")
        print(f"  Barcha fayllar bo'yicha o'quvchilar : {all_scores_count}")
        print(f"  Umumiy o'rtacha ball                : {grand_average:.2f}")
    else:
        print("  Hisoblash uchun ma'lumot topilmadi.")
    print(f"{'='*60}\n")


def main():
    # Fayl nomlarini argumentdan ol
    if len(sys.argv) > 1:
        filepaths = []
        for arg in sys.argv[1:]:
            # Glob pattern qo'llab-quvvatlash (masalan: *.xlsx)
            matched = glob.glob(arg)
            if matched:
                filepaths.extend(matched)
            elif os.path.exists(arg):
                filepaths.append(arg)
            else:
                print(f"[OGOHLANTIRISH] Fayl topilmadi: {arg}")
    else:
        # Joriy papkadan barcha Excel fayllarni ol
        filepaths = glob.glob("*.xlsx") + glob.glob("*.xls") + glob.glob("*.xlsm")
        if not filepaths:
            print("Excel fayllari topilmadi!")
            print("Ishlatish: python chsb_calculator.py fayl1.xlsx fayl2.xlsx ...")
            sys.exit(1)

    if not filepaths:
        print("Qayta ishlanadigan fayl yo'q.")
        sys.exit(1)

    print(f"\nTopilgan fayllar: {len(filepaths)} ta")
    for f in filepaths:
        print(f"  - {f}")

    all_results = []
    for filepath in sorted(filepaths):
        result = process_excel_file(filepath)
        all_results.append(result)

    # Bir nechta fayl bo'lsa umumiy xulosa chiqar
    if len(filepaths) > 1:
        print_summary(all_results)


if __name__ == "__main__":
    main()
